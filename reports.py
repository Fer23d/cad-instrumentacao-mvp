from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HEADERS = [
    "Seq.",
    "Tag",
    "Tipo",
    "Origem",
    "Bloco",
    "Layer",
    "X",
    "Y",
    "Conf.",
    "Obs.",
]


def create_excel_report(analysis: dict[str, Any], output_path: str | Path) -> Path:
    path = Path(output_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Instrumentos"

    sheet["A1"] = "Relatorio de Instrumentacao CAD"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet.merge_cells("A1:J1")

    project = analysis.get("project", {})
    sheet["A2"] = f"Projeto: {project.get('project_name') or '-'}"
    sheet["A3"] = f"Cliente: {project.get('client_name') or '-'}"
    sheet["A4"] = f"Responsavel tecnico: {project.get('technical_owner') or '-'}"
    sheet["A5"] = f"Tipo da planta: {project.get('drawing_type') or '-'}"
    sheet["A6"] = f"Arquivo: {analysis['filename']}"
    sheet["A7"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sheet["A8"] = f"Total identificado: {analysis['total_instruments']}"

    start_row = 10
    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="233142")
        cell.alignment = Alignment(horizontal="center")

    for row_index, instrument in enumerate(analysis["instruments"], start=start_row + 1):
        values = _instrument_row(instrument)
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=col, value=value)

    summary_sheet = workbook.create_sheet("Resumo")
    summary_sheet["A1"] = "Resumo por tipo"
    summary_sheet["A1"].font = Font(size=14, bold=True)
    summary_sheet.append(["Tipo", "Quantidade"])
    for instrument_type, total in analysis["summary_by_type"].items():
        summary_sheet.append([instrument_type, total])

    warnings_sheet = workbook.create_sheet("Alertas")
    warnings_sheet["A1"] = "Alertas"
    warnings_sheet["A1"].font = Font(size=14, bold=True)
    for warning in analysis["warnings"]:
        warnings_sheet.append([warning])

    project_sheet = workbook.create_sheet("Dados do Projeto")
    project_sheet.append(["Campo", "Valor"])
    project_sheet.append(["Projeto", project.get("project_name") or "-"])
    project_sheet.append(["Cliente", project.get("client_name") or "-"])
    project_sheet.append(["Responsavel tecnico", project.get("technical_owner") or "-"])
    project_sheet.append(["Tipo da planta", project.get("drawing_type") or "-"])
    project_sheet.append(["Arquivo", analysis["filename"]])
    project_sheet.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")])
    project_sheet.append(["Logo usado no PDF", "Sim" if project.get("logo_path") else "Nao"])

    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 45)

    workbook.save(path)
    return path


def create_pdf_report(analysis: dict[str, Any], output_path: str | Path) -> Path:
    path = Path(output_path)
    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        rightMargin=1.1 * cm,
        leftMargin=1.1 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm,
    )
    styles = getSampleStyleSheet()
    story = []
    project = analysis.get("project", {})

    story.append(Spacer(1, 1.2 * cm))
    logo = _logo_image(project.get("logo_path"))
    if logo is not None:
        story.append(logo)
        story.append(Spacer(1, 0.35 * cm))
    story.append(Paragraph("Relatorio de Instrumentacao CAD", styles["Title"]))
    story.append(Spacer(1, 0.55 * cm))
    cover_data = [
        ["Projeto", project.get("project_name") or "-"],
        ["Cliente", project.get("client_name") or "-"],
        ["Responsavel tecnico", project.get("technical_owner") or "-"],
        ["Tipo da planta", project.get("drawing_type") or "-"],
        ["Arquivo analisado", analysis["filename"]],
        ["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Instrumentos identificados", str(analysis["total_instruments"])],
    ]
    cover_table = Table(cover_data, colWidths=[5 * cm, 14 * cm])
    cover_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#233142")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b8c1cc")),
                ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#f5f7fa")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(cover_table)
    story.append(Spacer(1, 0.55 * cm))
    story.append(Paragraph("Resumo Executivo", styles["Heading2"]))
    story.append(
        Paragraph(
            "Este relatorio consolida os instrumentos identificados automaticamente na planta CAD, "
            "com classificacao por tipo, tag, layer, bloco de origem e coordenadas aproximadas.",
            styles["Normal"],
        )
    )
    story.append(PageBreak())
    story.append(Paragraph("Lista de Instrumentos", styles["Title"]))
    story.append(Spacer(1, 0.35 * cm))

    if analysis["summary_by_type"]:
        summary_data = [["Tipo", "Quantidade"], *analysis["summary_by_type"].items()]
        summary_table = Table(summary_data, colWidths=[8 * cm, 3 * cm])
        summary_table.setStyle(_table_style())
        story.append(summary_table)
        story.append(Spacer(1, 0.45 * cm))

    table_data = [HEADERS]
    for instrument in analysis["instruments"]:
        table_data.append([_pdf_cell(value, styles) for value in _instrument_row(instrument)])

    table = Table(table_data, repeatRows=1, colWidths=[1.0 * cm, 2.0 * cm, 4.0 * cm, 1.6 * cm, 3.0 * cm, 2.4 * cm, 1.7 * cm, 1.7 * cm, 1.5 * cm, 4.0 * cm])
    table.setStyle(_table_style())
    story.append(table)

    if analysis["warnings"]:
        story.append(Spacer(1, 0.45 * cm))
        story.append(Paragraph("Alertas", styles["Heading2"]))
        for warning in analysis["warnings"]:
            story.append(Paragraph(f"- {warning}", styles["Normal"]))

    doc.build(story)
    return path


def _instrument_row(instrument: dict[str, Any]) -> list[Any]:
    return [
        instrument["sequence"],
        instrument["tag"],
        instrument["instrument_type"],
        instrument["source"],
        instrument["block_name"],
        instrument["layer"],
        instrument["x"],
        instrument["y"],
        instrument["confidence"],
        instrument["notes"],
    ]


def _pdf_cell(value: Any, styles: Any) -> Paragraph:
    return Paragraph(str(value), styles["BodyText"])


def _logo_image(path_value: Any) -> Image | None:
    if not path_value:
        return None
    path = Path(str(path_value))
    if not path.exists():
        return None
    try:
        image = Image(str(path), width=5.2 * cm, height=2.4 * cm, kind="proportional")
        image.hAlign = "LEFT"
        return image
    except Exception:
        return None


def _table_style() -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#233142")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b8c1cc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
    )
